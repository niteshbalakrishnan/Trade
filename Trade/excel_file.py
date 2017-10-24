import os
import sys
sys.path.append("C:\Phython Projects\Packages\jdcal-1.3")
sys.path.append("C:\Phython Projects\Packages\et_xmlfile-1.0.1")
from openpyxl import load_workbook, Workbook
os.remove("\\MERCURY\Departments\MISC\hrdata\nitesh\CM.xlsx")
tmp = open("C:\Temp\CM.xlsx",'w')
fwb = Workbook()
fs = fwb.active
wb = load_workbook('C:\Temp\Summary_HC_12_NEW.xlsx')
ws1 = wb.active
wb = load_workbook('C:\Temp\Summary_HC_13_NEW.xlsx')
ws2 = wb.active
wb = load_workbook('C:\Temp\Summary_HC_14new.xlsx')
ws3 = wb.active
#sheet_to_copy = ws1.get_sheet_by_name("Sheet1")

fs = fwb.get_sheet_by_name("Sheet")
fs.title = 'first'
fwb.create_sheet("second")
fwb.create_sheet("third")
fs = fwb.get_sheet_by_name("first")
for i in range(1,101):
 for j in range(1,101):
     fs.cell(row=i, column=j).value= ws1.cell(row=i, column=j).value

fs = fwb.get_sheet_by_name("second")
for i in range(1,101):
 for j in range(1,101):
     fs.cell(row=i, column=j).value= ws2.cell(row=i, column=j).value

fs = fwb.get_sheet_by_name("third")
for i in range(1,101):
 for j in range(1,101):
     fs.cell(row=i, column=j).value= ws3.cell(row=i, column=j).value





fwb.save('C:\Temp\CM.xlsx')
print('dd')

